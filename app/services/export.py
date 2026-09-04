import csv
import io
from datetime import date

from openpyxl import Workbook
from sqlalchemy.orm import Session, joinedload

from app.models import Report, ReportLine, ReportStatus
from app.services.labels import status_label


def _fmt_date(d: date | None) -> str:
    return d.strftime("%d.%m.%Y") if d else ""


def _reports_query(db: Session, from_date: date | None, to_date: date | None, status: ReportStatus | None):
    q = db.query(Report).options(
        joinedload(Report.employee),
        joinedload(Report.lines).joinedload(ReportLine.work_item),
    )
    if from_date:
        q = q.filter(Report.report_date >= from_date)
    if to_date:
        q = q.filter(Report.report_date <= to_date)
    if status:
        q = q.filter(Report.status == status)
    return q.order_by(Report.report_date.desc(), Report.id.desc()).all()


def export_txt(db: Session, from_date: date | None = None, to_date: date | None = None) -> str:
    reports = _reports_query(db, from_date, to_date, ReportStatus.approved)
    if not reports:
        reports = _reports_query(db, from_date, to_date, ReportStatus.submitted)
    if not reports:
        return "Нет данных за выбранный период."

    lines_out = [f"Отчёты за период: {_fmt_date(from_date) or '…'} — {_fmt_date(to_date) or '…'}", ""]
    for report in reports:
        lines_out.append(f"=== {report.report_date.strftime('%d.%m.%Y')} | {report.employee.full_name} ===")
        lines_out.append(f"Статус: {status_label(report.status)}")
        for line in report.lines:
            w = line.work_item
            note = f" ({line.note})" if line.note else ""
            lines_out.append(
                f"  {w.code} | {w.name} | {line.quantity} {w.unit} × {line.unit_price} = {line.amount} руб.{note}"
            )
        lines_out.append(f"ИТОГО: {report.total_amount:.2f} руб.\n")
    return "\n".join(lines_out)


def export_csv(db: Session, from_date: date | None = None, to_date: date | None = None) -> str:
    reports = _reports_query(db, from_date, to_date, None)
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([
        "Дата", "ФИО", "Город", "Статус", "Код", "Работа",
        "Кол-во", "Ед. изм.", "Цена, руб.", "Сумма, руб.", "Примечание",
    ])
    for report in reports:
        for line in report.lines:
            w = line.work_item
            writer.writerow([
                report.report_date.strftime("%d.%m.%Y"),
                report.employee.full_name,
                report.employee.city,
                status_label(report.status),
                w.code,
                w.name,
                line.quantity,
                w.unit,
                line.unit_price,
                line.amount,
                line.note or "",
            ])
    return buf.getvalue()


def export_xlsx(db: Session, from_date: date | None = None, to_date: date | None = None) -> bytes:
    reports = _reports_query(db, from_date, to_date, None)
    wb = Workbook()

    ws_detail = wb.active
    ws_detail.title = "Детализация"
    ws_detail.append([
        "Дата", "ФИО", "Город", "Статус", "Код", "Наименование работы",
        "Кол-во", "Ед. изм.", "Цена, руб.", "Сумма, руб.", "Примечание",
    ])
    for report in reports:
        for line in report.lines:
            w = line.work_item
            ws_detail.append([
                report.report_date.strftime("%d.%m.%Y"),
                report.employee.full_name,
                report.employee.city,
                status_label(report.status),
                w.code,
                w.name,
                line.quantity,
                w.unit,
                line.unit_price,
                line.amount,
                line.note or "",
            ])

    ws_summary = wb.create_sheet("Сводка по ФИО")
    ws_summary.append(["ФИО", "Город", "Период с", "Период по", "Сумма, руб."])
    totals: dict[tuple[str, str], float] = {}
    for report in reports:
        key = (report.employee.full_name, report.employee.city)
        totals[key] = totals.get(key, 0) + report.total_amount
    for (fio, city), total in sorted(totals.items()):
        ws_summary.append([fio, city, _fmt_date(from_date), _fmt_date(to_date), round(total, 2)])

    ws_daily = wb.create_sheet("По дням")
    ws_daily.append(["Дата", "ФИО", "Сумма, руб.", "Статус"])
    for report in reports:
        ws_daily.append([
            report.report_date.strftime("%d.%m.%Y"),
            report.employee.full_name,
            report.total_amount,
            status_label(report.status),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
