import json
from pathlib import Path

from sqlalchemy.orm import Session

from app.config import settings
from app.models import Employee, Report, ReportLine, UserRole, WorkItem


def load_json(name: str) -> list[dict]:
    path = settings.data_dir / name
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _load_work_items_from_json(db: Session) -> tuple[int, int]:
    """Загрузить прайс из seed_works.json и internal_tariffs.json."""
    works_count = 0
    internal_count = 0
    seen_codes: set[str] = set()
    for item in load_json("seed_works.json"):
        code = item["code"]
        if code in seen_codes:
            continue
        seen_codes.add(code)
        db.add(
            WorkItem(
                code=item["code"],
                name=item["name"],
                unit=item.get("unit") or "шт.",
                tariff=item.get("tariff"),
                operation_price=item.get("operation_price"),
                category=item.get("category") or "Прочее",
                is_internal=False,
            )
        )
        works_count += 1
    for item in load_json("internal_tariffs.json"):
        db.add(
            WorkItem(
                code=item["code"],
                name=item["name"],
                unit=item.get("unit") or "шт.",
                tariff=item.get("tariff"),
                operation_price=item.get("tariff"),
                category=item.get("category") or "Внутренние",
                is_internal=True,
            )
        )
        internal_count += 1
    return works_count, internal_count


def seed_database(db: Session) -> dict:
    stats = {"works": 0, "employees": 0, "internal": 0}

    if db.query(WorkItem).count() == 0:
        stats["works"], stats["internal"] = _load_work_items_from_json(db)

    if db.query(Employee).count() == 0:
        stats["employees"] = _seed_employees(db)
    else:
        stats["employees"] = ensure_service_accounts(db)

    db.commit()
    return stats


def _seed_service_accounts(db: Session) -> int:
    """Диспетчер и администратор (системные учётки)."""
    added = 0
    service = [
        ("Диспетчер", UserRole.dispatcher, "0000"),
        ("Администратор", UserRole.admin, "admin"),
    ]
    for full_name, role, pin in service:
        if not db.query(Employee).filter(Employee.full_name == full_name).first():
            db.add(Employee(full_name=full_name, city="Офис", role=role, pin=pin))
            added += 1
    return added


def _seed_employees(db: Session) -> int:
    count = 0
    for item in load_json("seed_employees.json"):
        db.add(
            Employee(
                full_name=item["full_name"],
                city=item.get("city", "Город"),
                role=UserRole(item.get("role", "master")),
                pin=item.get("pin", "1234"),
            )
        )
        count += 1
    count += _seed_service_accounts(db)
    return count


def ensure_service_accounts(db: Session) -> int:
    """Добавить диспетчера и админа, если их нет."""
    added = _seed_service_accounts(db)
    if added:
        db.commit()
    return added


def reset_employees(db: Session) -> int:
    """Удалить всех сотрудников и отчёты, загрузить из seed_employees.json."""
    db.query(ReportLine).delete()
    db.query(Report).delete()
    db.query(Employee).delete()
    db.flush()
    count = _seed_employees(db)
    db.commit()
    return count


def reset_works(db: Session) -> tuple[int, int]:
    """Удалить прайс и отчёты, загрузить демо-прайс из JSON."""
    db.query(ReportLine).delete()
    db.query(Report).delete()
    db.query(WorkItem).delete()
    db.flush()
    works, internal = _load_work_items_from_json(db)
    db.commit()
    return works, internal


def import_price_from_xlsx(db: Session, file_path: Path) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    wsd = wb["Данные"]
    count = 0
    for row in range(4, wsd.max_row + 1):
        code = wsd.cell(row, 1).value
        name = wsd.cell(row, 2).value
        unit = wsd.cell(row, 5).value
        tariff = wsd.cell(row, 7).value
        op_price = wsd.cell(row, 11).value
        if not name or not code or "." not in str(code):
            continue
        code_s = str(code).strip()
        existing = db.query(WorkItem).filter(WorkItem.code == code_s).first()
        tariff_f = float(tariff) if isinstance(tariff, (int, float)) else None
        op_f = float(op_price) if isinstance(op_price, (int, float)) else None
        cat_code = code_s.split(".")[0]
        category = f"Раздел {cat_code}"
        if existing:
            existing.name = str(name).strip()
            existing.unit = str(unit).strip() if unit else "шт."
            existing.tariff = tariff_f
            existing.operation_price = op_f
            existing.category = category
        else:
            db.add(
                WorkItem(
                    code=code_s,
                    name=str(name).strip(),
                    unit=str(unit).strip() if unit else "шт.",
                    tariff=tariff_f,
                    operation_price=op_f,
                    category=category,
                )
            )
        count += 1
    db.commit()
    return count
